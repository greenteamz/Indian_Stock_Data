import yfinance as yf
from datetime import datetime, timedelta
import gspread
from google.oauth2.service_account import Credentials
from google.cloud import bigquery
from google.api_core.exceptions import NotFound
from openpyxl import Workbook, load_workbook
import csv
import os
import pytz
import time
import logging
import pandas as pd
from collections import defaultdict
import sys

# Define IST timezone
IST = pytz.timezone('Asia/Kolkata')

# Current IST datetime
ist_now = datetime.now(IST)

# Extract the date part from IST datetime
ist_date = ist_now.date()

current_year = ist_date.year  # Extract the current year

# Generate log and CSV file names 
log_filename = f"log_NSE_Daily_{ist_now.strftime('%Y-%m-%d_%H-%M-%S')}.txt"
master_log_filename = f"Log_Master_NSE_{current_year}.txt"
csv_filename = f"NSE_Stock_Master_{current_year}.csv"  # Append data for the same day
csv_filename_daily = f"NSE_Stock_Daily_{ist_now.strftime('%Y-%m-%d_%H-%M-%S')}.csv"  # Append data for the same day
excel_filename = f"NSE_Stock_Master_All_{current_year}.xlsx"  # Excel file for today

# Define base directory
BASE_DIR = f"NSE_{current_year}"

# Subdirectories under NSE
MASTER_DIR = os.path.join(BASE_DIR, "master_nse")
LOGS_DIR = os.path.join(BASE_DIR, "logs_nse")
CSV_DIR = os.path.join(BASE_DIR, "csv_nse")

# Paths for logs, CSV, and Excel
MASTER_LOG_FILE_PATH = os.path.join(MASTER_DIR, master_log_filename)
LOG_FILE_PATH = os.path.join(LOGS_DIR, log_filename)
MASTER_CSV_FILE_PATH = os.path.join(MASTER_DIR, csv_filename)
Daily_CSV_FILE_PATH  = os.path.join(CSV_DIR, csv_filename_daily)
EXCEL_FILE_PATH = os.path.join(MASTER_DIR, excel_filename)

# Ensure all required directories exist
os.makedirs(MASTER_DIR, exist_ok=True)
os.makedirs(LOGS_DIR, exist_ok=True)
os.makedirs(CSV_DIR, exist_ok=True)

# Log function
def log_message(message):
    """Log messages to a file and print to console."""
    timestamp = datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")
    with open(LOG_FILE_PATH, "a") as log_file:
        log_file.write(f"[{timestamp}] {message}\n")
    with open(MASTER_LOG_FILE_PATH, "a") as master_log_file:
        master_log_file.write(f"[{timestamp}] {message}\n")
    print(f"[{timestamp}] {message}")

# Authenticate using the same service_account.json for both BigQuery and Google Sheets
SERVICE_ACCOUNT_FILE = "service_account.json"

# Google Sheets authentication
gc = gspread.service_account(filename=SERVICE_ACCOUNT_FILE)


# Open Google Spreadsheet
spreadsheet = gc.open('NSE_symbol')  # Replace with your Google Sheet name
#source_worksheet = spreadsheet.worksheet('symbol')  # Replace with your sheet name
source_worksheet = spreadsheet.worksheet('symbol')  # Test sheet name

# Fetch all stock symbols from the first column
symbols = source_worksheet.col_values(1)[1:]  # Skip header row
symbols = [symbol if symbol.endswith('.NS') else f"{symbol}.NS" for symbol in symbols]

# Define BigQuery dataset and table with the project ID
PROJECT_ID = "stockautomation-442015"  # Replace with your project ID
BQ_DATASET = f"nse_stock_{current_year}"  # Replace with your dataset name
BQ_TABLE = f"{PROJECT_ID}.{BQ_DATASET}.daily_nse_stock_{current_year}"  # Fully-qualified table name

# Define schema for BigQuery table
headers = [
    "fullTimeEmployees", "auditRisk", "boardRisk",
    "compensationRisk", "shareHolderRightsRisk", "overallRisk", "maxAge", "priceHint",
    "regularMarketOpen", "regularMarketDayLow", "regularMarketDayHigh", "dividendRate",
    "dividendYield", "exDividendDate", "payoutRatio", "fiveYearAvgDividendYield", "beta",
    "volume", "regularMarketVolume", "averageVolume",
    "returnOnAssets", "returnOnEquity", "freeCashflow", "operatingCashflow",
    "twoHundredDayAverage", "trailingAnnualDividendRate", "trailingAnnualDividendYield",
    "enterpriseValue", "floatShares", "sharesOutstanding",
    "heldPercentInsiders", "heldPercentInstitutions", "impliedSharesOutstanding",
    "bookValue", "priceToBook", "earningsQuarterlyGrowth", "trailingEps", "forwardEps",
    "52WeekChange", "lastDividendValue", "lastDividendDate", "exchange", "quoteType", 
    "totalCash", "totalCashPerShare", "ebitda", "totalDebt",
    "averageVolume10days", "averageDailyVolume10Day", "marketCap", "fiftyTwoWeekLow",
    "fiftyTwoWeekHigh", "priceToSalesTrailing12Months", "fiftyDayAverage",
    "quickRatio", "currentRatio", "totalRevenue", "debtToEquity", "revenuePerShare",
    "earningsGrowth", "revenueGrowth", "grossMargins", "ebitdaMargins", "operatingMargins",  "profitMargins",
    "previousClose", "open", "dayLow", "dayHigh", "regularMarketPreviousClose", "trailingPE", "forwardPE",
    "symbol", "shortName", "longName",  "industry", "sector", "currentPrice", "targetHighPrice", "targetLowPrice",
    "targetMeanPrice", "targetMedianPrice", "recommendationMean", "recommendationKey",
    "numberOfAnalystOpinions"
]

# Define a data type mapping for headers
data_type_map = {
    "industry": "STRING",
    "sector": "STRING",
    "fullTimeEmployees": "FLOAT",  # Integer field
    "auditRisk": "FLOAT",
    "boardRisk": "FLOAT",
    "compensationRisk": "FLOAT",
    "shareHolderRightsRisk": "FLOAT",
    "overallRisk": "FLOAT",
    "maxAge": "FLOAT",
    "priceHint": "FLOAT",
    "previousClose": "FLOAT",
    "open": "FLOAT",
    "dayLow": "FLOAT",
    "dayHigh": "FLOAT",
    "regularMarketPreviousClose": "FLOAT",
    "regularMarketOpen": "FLOAT",
    "regularMarketDayLow": "FLOAT",
    "regularMarketDayHigh": "FLOAT",
    "dividendRate": "FLOAT",
    "dividendYield": "FLOAT",
    "exDividendDate": "DATE",
    "payoutRatio": "FLOAT",
    "fiveYearAvgDividendYield": "FLOAT",
    "beta": "FLOAT",
    "trailingPE": "FLOAT",
    "forwardPE": "FLOAT",
    "volume": "FLOAT",
    "regularMarketVolume": "FLOAT",
    "averageVolume": "FLOAT",
    "averageVolume10days": "FLOAT",
    "averageDailyVolume10Day": "FLOAT",
    "marketCap": "FLOAT",
    "fiftyTwoWeekLow": "FLOAT",
    "fiftyTwoWeekHigh": "FLOAT",
    "priceToSalesTrailing12Months": "FLOAT",
    "fiftyDayAverage": "FLOAT",
    "twoHundredDayAverage": "FLOAT",
    "trailingAnnualDividendRate": "FLOAT",
    "trailingAnnualDividendYield": "FLOAT",
    "enterpriseValue": "FLOAT",
    "profitMargins": "FLOAT",
    "floatShares": "FLOAT",
    "sharesOutstanding": "FLOAT",
    "heldPercentInsiders": "FLOAT",
    "heldPercentInstitutions": "FLOAT",
    "impliedSharesOutstanding": "FLOAT",
    "bookValue": "FLOAT",
    "priceToBook": "FLOAT",
    "earningsQuarterlyGrowth": "FLOAT",
    "trailingEps": "FLOAT",
    "forwardEps": "FLOAT",
    "52WeekChange": "FLOAT",
    "lastDividendValue": "FLOAT",
    "lastDividendDate": "DATE",
    "exchange": "STRING",
    "quoteType": "STRING",
    "symbol": "STRING",
    "shortName": "STRING",
    "longName": "STRING",
    "currentPrice": "FLOAT",
    "targetHighPrice": "FLOAT",
    "targetLowPrice": "FLOAT",
    "targetMeanPrice": "FLOAT",
    "targetMedianPrice": "FLOAT",
    "recommendationMean": "FLOAT",
    "recommendationKey": "STRING",
    "numberOfAnalystOpinions": "FLOAT",
    "totalCash": "FLOAT",
    "totalCashPerShare": "FLOAT",
    "ebitda": "FLOAT",
    "totalDebt": "FLOAT",
    "quickRatio": "FLOAT",
    "currentRatio": "FLOAT",
    "totalRevenue": "FLOAT",
    "debtToEquity": "FLOAT",
    "revenuePerShare": "FLOAT",
    "returnOnAssets": "FLOAT",
    "returnOnEquity": "FLOAT",
    "freeCashflow": "FLOAT",
    "operatingCashflow": "FLOAT",
    "earningsGrowth": "FLOAT",
    "revenueGrowth": "FLOAT",
    "grossMargins": "FLOAT",
    "ebitdaMargins": "FLOAT",
    "operatingMargins": "FLOAT",
    "Today_Growth": "FLOAT",
    "Calculated_Score": "FLOAT",
    "Score_Recommendation": "STRING",
    "Conservative_Invs_Recom": "STRING",
    "Conservative_Invs_Reson": "STRING",
    "Growth_Invs_Recom": "STRING",
    "Growth_Invs_Reson": "STRING",
    "Momentum_Invs_Recom": "STRING",
    "Momentum_Invs_Reson": "STRING",
    "sector_rank": "FLOAT",
    "industry_rank": "FLOAT",
}

rank_headers = ["sector_rank", "industry_rank"]
ROW_COUNTER_FILE = os.path.join(MASTER_DIR, "nse_row_counter.txt")

current_year = datetime.now().year
consolidated_file = f"Consolidated_{current_year}.csv"
consolidated_file_path = os.path.join(MASTER_DIR, consolidated_file)
    
# Initialize row_insert_order
def initialize_row_counter():
    if not os.path.exists(ROW_COUNTER_FILE):
        with open(ROW_COUNTER_FILE, "w") as f:
            f.write("1")  # Start counter at 1

def get_current_row_counter():
    with open(ROW_COUNTER_FILE, "r") as f:
        return int(f.read().strip())

def update_row_counter(new_value):
    with open(ROW_COUNTER_FILE, "w") as f:
        f.write(str(new_value))

# Initialize the row counter if not already done
initialize_row_counter()


# Add "Previous Day Date" to headers
# PREVIOUS_DAY_DATE = (ist_date - timedelta(days=1)).strftime('%Y-%m-%d') ist_now.strftime('%Y-%m-%d_%H-%M-%S')
PREVIOUS_DAY_DATETIME = ist_now.strftime('%Y-%m-%d %H:%M:%S')
#headers_with_date = ["row_insert_order", "PreviousDayDate", "Symbol_Input"] + headers

score_headers = ["Today_Growth", "Calculated_Score", "Score_Recommendation", "Conservative_Invs_Recom", "Conservative_Invs_Reson", "Growth_Invs_Recom", "Growth_Invs_Reson", "Momentum_Invs_Recom", "Momentum_Invs_Reson"]

def ensure_dataset_exists():
    try:
        bq_client.get_dataset(BQ_DATASET)
        log_message(f"Dataset '{BQ_DATASET}' exists.")
    except NotFound:
        dataset = bigquery.Dataset(f"{PROJECT_ID}.{BQ_DATASET}")
        bq_client.create_dataset(dataset)
        log_message(f"Created dataset '{BQ_DATASET}'.")

def ensure_table_exists():
    try:
        # Check if the table already exists
        table = bq_client.get_table(BQ_TABLE)
        log_message(f"Table '{BQ_TABLE}' already exists.")
    except NotFound:
        # Table does not exist, create it
        # Build the schema dynamically
        schema = [bigquery.SchemaField("row_insert_order", "INTEGER"), bigquery.SchemaField("PreviousDayDate", "DATETIME"), bigquery.SchemaField("Symbol_Input", "STRING"),] + [
                bigquery.SchemaField(header, data_type_map.get(header, "STRING"))
                for header in headers
                ] + [ bigquery.SchemaField(header, data_type_map.get(header, "STRING")) for header in score_headers ] +[ 
                bigquery.SchemaField(header, data_type_map.get(header, "STRING")) for header in rank_headers ]
        
        table = bigquery.Table(BQ_TABLE, schema=schema)
        bq_client.create_table(table)
        log_message(f"Created table '{BQ_TABLE}'.")
    except Exception as e:
        log_message(f"Error ensuring table exists: {e}")

def create_consolidated_csv():
    """
    Creates a new consolidated CSV file and appends data from the provided DataFrame.

    Args:
        df: pandas DataFrame containing stock data.

    Returns:
        None
    """
    log_message("Started create_consolidated_csv.")
    df = pd.read_csv(Daily_CSV_FILE_PATH)


    # Create headers dynamically
    today = datetime.now()
    end_of_year = datetime(current_year, 12, 31)
    date_range = pd.date_range(start=datetime(current_year, 1, 1), end=end_of_year)
    business_days = date_range[date_range.weekday < 5]  # Filter out weekends

    # Create header names for daily, weekly, monthly, and yearly changes
    daily_change_headers = [f"D{day.strftime('%d_%m')}_Diff" for day in business_days]

    # Create weekly change headers 
    weekly_change_headers = []
    current_week = 1
    current_month = 1
    for i in range(0, len(business_days), 5):  # Iterate by 5 business days (assuming a 5-day workweek)
        week_start = business_days[i]
        week_end = business_days[min(i + 4, len(business_days) - 1)]  # Handle potential incomplete weeks at the end of the year
        weekly_change_headers.append(f"W{current_week:02d}_{week_start.month:02d}") 
        current_week += 1
        if week_end.month > week_start.month:
            current_month += 1 
            
    monthly_change_headers = [f"M{month:02d}_{year}" for year in range(current_year, current_year + 1) for month in range(1, 13)]
    yearly_change_headers = [str(current_year)]

    # Add new headers to the list
    headers = list(df.columns) + ["Stock_Volatile", "Stock_Volatile_Percentage" ] # Include all existing columns and "Stock_Volatile" first
    headers.extend(daily_change_headers)
    headers.extend(weekly_change_headers)
    headers.extend(monthly_change_headers)
    headers.extend(yearly_change_headers)
    headers.extend(["Expert_Comments", "Expert_Review"])

    # Create the consolidated CSV file
    if not os.path.exists(consolidated_file_path):
        with open(consolidated_file_path, mode="w", newline="") as csv_file:
            writer = csv.writer(csv_file)
            writer.writerow(headers)
            log_message(f"Created consolidated_file_path= '{consolidated_file_path}'.")

    # Append data to the consolidated CSV file
    with open(consolidated_file_path, mode="a", newline="") as csv_file:
        writer = csv.writer(csv_file)
        for index, row in df.iterrows():
            #data_row = [row[col] for col in headers]  # Create a list with all data
            # Create a list with all data, handling potential absence of "Stock_Volatile"
            data_row = [row[col] if col in df.columns else None for col in headers]
            writer.writerow(data_row)
    log_message(f"Append data to '{consolidated_file_path}' file.")

    # Update daily change data (only for today's date)
    today_str = today.strftime("%d_%m")
    today_change_header = f"D{today_str}_Diff"
    if today_change_header in daily_change_headers:
        update_consolidated_data(consolidated_file_path, "Today_Growth", [today_change_header])

    # Calculate and update weekly, monthly, and yearly changes
    calculate_and_update_changes(consolidated_file_path, daily_change_headers, weekly_change_headers, monthly_change_headers, business_days, yearly_change_headers)

    # Calculate and add "Stock_Volatile" column
    calculate_stock_volatility(consolidated_file_path)
    
    def load_consolidated_csv_to_bigquery():
        """Load consolidated CSV data into a new BigQuery dataset and table."""
        try:
            # Define new BigQuery dataset and table for consolidated data
            consolidated_dataset = f"consolidated_{current_year}"
            consolidated_table = f"{PROJECT_ID}.{consolidated_dataset}.consolidated_nse_stock_{current_year}"

            # Ensure the new dataset exists
            try:
                bq_client.get_dataset(consolidated_dataset)
                log_message(f"Dataset '{consolidated_dataset}' exists.")
            except NotFound:
                dataset = bigquery.Dataset(f"{PROJECT_ID}.{consolidated_dataset}")
                bq_client.create_dataset(dataset)
                log_message(f"Created dataset '{consolidated_dataset}'.")

            # Define schema for the consolidated table
            consolidated_schema = [
                bigquery.SchemaField(header, data_type_map.get(header, "STRING"))
                for header in headers
            ]             # + [
            #     bigquery.SchemaField("Stock_Volatile", "FLOAT"),
            #     bigquery.SchemaField("Stock_Volatile_Percentage", "FLOAT"),
            #     bigquery.SchemaField("Expert_Comments", "STRING"),
            #     bigquery.SchemaField("Expert_Review", "STRING")
            # ] + [
            #     bigquery.SchemaField(header, "FLOAT") for header in daily_change_headers
            # ] + [
            #     bigquery.SchemaField(header, "FLOAT") for header in weekly_change_headers
            # ] + [
            #     bigquery.SchemaField(header, "FLOAT") for header in monthly_change_headers
            # ] + [
            #     bigquery.SchemaField(header, "FLOAT") for header in yearly_change_headers
            # ]

            # Create the consolidated table if it doesn't exist
            try:
                bq_client.get_table(consolidated_table)
                log_message(f"Table '{consolidated_table}' already exists.")
            except NotFound:
                table = bigquery.Table(consolidated_table, schema=consolidated_schema)
                bq_client.create_table(table)
                log_message(f"Created table '{consolidated_table}'.")

            # Preprocess the consolidated CSV file
            processed_data = preprocess_data(consolidated_file_path)
            
            # Write processed data back to a temporary CSV for BigQuery loading
            consolidated_tmp_path = "consolidated_tmp.csv"
            
            # Check if the file exists, and delete it if it does
            if os.path.exists(consolidated_tmp_path):
                os.remove(consolidated_tmp_path)
                log_message(f"Deleted the file before start - {consolidated_tmp_path}.")
                
            with open(consolidated_tmp_path, "w", newline="") as temp_csv:
                writer = csv.DictWriter(temp_csv, fieldnames=processed_data[0].keys())
                writer.writeheader()  # Write headers
                writer.writerows(processed_data)  # Write processed rows
            
            log_message(f"Start to load data to BigQuery from {consolidated_tmp_path}.")

            # Load the processed data into BigQuery
            with open(consolidated_tmp_path, "rb") as csv_file:
                job_config = bigquery.LoadJobConfig(
                source_format=bigquery.SourceFormat.CSV,
                skip_leading_rows=1,  # Skip header row
                write_disposition="WRITE_APPEND",  # Append data
                autodetect=False,
                max_bad_records=500,  # Tolerate up to 500 bad rows
                )
                load_job = bq_client.load_table_from_file(
                csv_file, consolidated_table, job_config=job_config
                )
                load_job.result()  # Wait for the job to complete
                log_message(f"Data loaded to BigQuery table '{consolidated_table}' from '{consolidated_tmp_path}'.")

                # Check for errors
                if load_job.errors:
                    log_message(f"Errors encountered during loading: {load_job.errors}")
                else:
                    log_message("Data loaded successfully, no errors.")
        except Exception as e:
            log_message(f"Error loading consolidated data to BigQuery: {e}")

    # Load the consolidated CSV data into the new BigQuery table
    load_consolidated_csv_to_bigquery()

def update_consolidated_data(file_path, source_column, target_columns):
    """
    Updates data in the consolidated CSV file by copying values from a source column to target columns.

    Args:
        file_path: Path to the consolidated CSV file.
        source_column: Name of the source column.
        target_columns: List of target column names.
    """
    
    df = pd.read_csv(file_path)
    for target_column in target_columns:
        df[target_column] = df[source_column]
    df.to_csv(file_path, index=False)
    log_message(f"Updated today change from '{source_column}' to {target_columns} .")

def calculate_and_update_changes(file_path, daily_change_headers, weekly_change_headers, monthly_change_headers, business_days, yearly_change_headers):
    """
    Calculates and updates weekly, monthly, and yearly changes in the consolidated CSV file.

    Args:
        file_path: Path to the consolidated CSV file.
        daily_change_headers: List of daily change headers.
        weekly_change_headers: List of weekly change headers.
        monthly_change_headers: List of monthly change headers.
    """
    df = pd.read_csv(file_path)

    # Calculate weekly changes
    for week_start_index in range(0, len(daily_change_headers), 5):  # Assuming 5 business days per week
        week_end_index = min(week_start_index + 4, len(daily_change_headers) - 1)  # Handle potential incomplete weeks
        week_change_header = weekly_change_headers[week_start_index // 5]  # Get the corresponding weekly header
        df[week_change_header] = df[daily_change_headers[week_start_index:week_end_index + 1]].sum(axis=1)


    # Calculate monthly changes based on daily changes
    for month in range(1, 13):
        month_str = f"M{month:02d}_{current_year}"
        month_change_header = f"{month_str}" 
        month_start = datetime(current_year, month, 1)
        month_end = month_start + pd.offsets.MonthEnd(0)
        business_days_in_month = business_days[(business_days >= month_start) & (business_days <= month_end)]
        daily_change_headers_in_month = [f"D{day.strftime('%d_%m')}_Diff" for day in business_days_in_month]
        df[month_change_header] = df[daily_change_headers_in_month].sum(axis=1)

    # Calculate yearly changes
    df[yearly_change_headers[0]] = df[monthly_change_headers].sum(axis=1)
    log_message(f"Updated today changes to weekly, monthly and yearly headers.")
    df.to_csv(file_path, index=False)

def calculate_stock_volatility(file_path):
    """
    Calculates and adds the "Stock_Volatile" column to the consolidated CSV file.

    Args:
        file_path: Path to the consolidated CSV file. Stock_Volatile_Percentage .round(2)
    """
    df = pd.read_csv(file_path)
    df["Stock_Volatile"] = (df["dayHigh"] - df["dayLow"])
    df["Stock_Volatile"] = df["Stock_Volatile"].round(2)
    df["Stock_Volatile_Percentage"] = (df["dayHigh"] - df["dayLow"]) / df["dayHigh"] * 100
    df["Stock_Volatile_Percentage"] = df["Stock_Volatile_Percentage"].round(2)
    df.to_csv(file_path, index=False)
    log_message(f"Updated stock volatile into {file_path}.")


def calculate_ranks(df, group_column, score_column, rank_column_name):
    """
    Calculate ranks within a group based on the score and append the rank as a new column.
    Handles NaN values in both group and score columns.
    """
    # Replace NaN values in score_column with a placeholder (-1) for rank calculation
    df[score_column] = pd.to_numeric(df[score_column], errors="coerce").fillna(-1)

    # Exclude rows where group_column is NaN
    valid_rows = ~df[group_column].isna()
    #print(df[group_column].isna())
    # Initialize the rank column with 0 as the default placeholder
    df[rank_column_name] = 0

    # Perform rank calculation only for valid rows
    df.loc[valid_rows, rank_column_name] = (
        df[valid_rows]
        .groupby(group_column)[score_column]
        .rank(ascending=False, method='dense', na_option='bottom')
        .fillna(0)
        .astype(int)
    )

    # Convert rank to integer (valid ranks) while keeping placeholder as 0
    df[rank_column_name] = df[rank_column_name].fillna(0).astype(int)
    log_message(f"Calculated ranks for '{group_column}' based on '{score_column}' and saved as '{rank_column_name}'")
    return df

    
def append_to_csv(data_row, total_symbol):
    """Append a row of data to the CSV file, adding the header only if it's a new file."""
    write_header = not os.path.exists(MASTER_CSV_FILE_PATH)  # Check if file exists

    with open(MASTER_CSV_FILE_PATH, mode="a", newline="") as csv_file:
        writer = csv.writer(csv_file)
        if write_header:
            writer.writerow(["row_insert_order", "PreviousDayDate", "Symbol_Input"] + headers + score_headers)  # Add header row
            log_message(f"Header added to CSV file: {MASTER_CSV_FILE_PATH}")
        writer.writerow(data_row)
        log_message(f"Appended data to Master CSV file: {MASTER_CSV_FILE_PATH}")

        log_message(f" count: {processed_count}/{total_symbol}")
        
        # If it's the last row, calculate the ranks and update the file
        if processed_count==total_symbol:
            # Load the CSV file into DataFrame to calculate ranks
            df = pd.read_csv(MASTER_CSV_FILE_PATH)
            df.columns = df.columns.str.strip()
            # Ensure 'sector' and 'industry' columns exist, adjust accordingly to your file's structure
            if 'sector' in df.columns and 'industry' in df.columns and 'Calculated_Score' in df.columns:
                # Calculate ranks for sector and industry based on 'Calculated_Score' column
                log_message(f"Required columns found in Master CSV, starting rank calculation")

                # Handle NaN values in required columns
                df['Calculated_Score'] = df['Calculated_Score'].fillna(-1)  # Replace NaN in score_column with -1
                df['sector'] = df['sector'].fillna('Unknown')  # Replace NaN in sector with 'Unknown'
                df['industry'] = df['industry'].fillna('Unknown')  # Replace NaN in industry with 'Unknown'

                df = calculate_ranks(df, 'sector', 'Calculated_Score', 'sector_rank')
                df = calculate_ranks(df, 'industry', 'Calculated_Score', 'industry_rank')

                # Save the updated DataFrame back to the same CSV file, overwriting it
                df.to_csv(MASTER_CSV_FILE_PATH, index=False)
                log_message(f"Sector and Industry Rank calculation completed and saved to Master CSV file: {MASTER_CSV_FILE_PATH}")
            else:
                print(df.columns)
                
    """Append a row of data to the CSV file, adding the header only if it's a new file."""
    write_header = not os.path.exists(Daily_CSV_FILE_PATH)  # Check if file exists

    with open(Daily_CSV_FILE_PATH, mode="a", newline="") as csv_file:
        writer = csv.writer(csv_file)
        if write_header:
            writer.writerow(["row_insert_order", "PreviousDayDate", "Symbol_Input"] + headers + score_headers)  # Add header row
            log_message(f"Header added to CSV file: {Daily_CSV_FILE_PATH}")
        writer.writerow(data_row)
        log_message(f"Appended data to Daily CSV file: {Daily_CSV_FILE_PATH}")

        # If it's the last row, calculate the ranks and update the file
        if processed_count==total_symbol:
            # Load the CSV file into DataFrame to calculate ranks
            df = pd.read_csv(Daily_CSV_FILE_PATH)
            # Ensure 'sector' and 'industry' columns exist, adjust accordingly to your file's structure
            if 'sector' in df.columns and 'industry' in df.columns and 'Calculated_Score' in df.columns:
                log_message(f"Required columns found in Daily CSV, starting rank calculation")
                
                # Handle NaN values in required columns
                df['Calculated_Score'] = df['Calculated_Score'].fillna(-1)  # Replace NaN in score_column with -1
                df['sector'] = df['sector'].fillna('Unknown')  # Replace NaN in sector with 'Unknown'
                df['industry'] = df['industry'].fillna('Unknown')  # Replace NaN in industry with 'Unknown'
                
                # Calculate ranks for sector and industry based on 'Calculated_Score' column sector_rank industry_rank
                df = calculate_ranks(df, 'sector', 'Calculated_Score', 'sector_rank')
                df = calculate_ranks(df, 'industry', 'Calculated_Score', 'industry_rank')

                # Save the updated DataFrame back to the same CSV file, overwriting it
                df.to_csv(Daily_CSV_FILE_PATH, index=False)
                log_message(f"Sector and Industry Rank calculation completed and saved to Daily CSV file: {Daily_CSV_FILE_PATH}")
                
            # Load or create the Excel workbook
            if os.path.exists(EXCEL_FILE_PATH):
                workbook = load_workbook(EXCEL_FILE_PATH)
                log_message(f"Loaded existing Excel file. {EXCEL_FILE_PATH}")
            else:
                workbook = Workbook()
                workbook.remove(workbook.active)  # Remove default sheet
                log_message(f"Created new Excel file. {EXCEL_FILE_PATH}")
    
            # Check if sheet already exists, create if not
            sheet_name = f"NSE_{ist_date}"
            if sheet_name not in workbook.sheetnames:
                # Create a new sheet if it doesn't exist
                workbook.create_sheet(sheet_name)
                sheet = workbook[sheet_name]
                sheet.append(df.columns.tolist())  # Add headers
                log_message(f"New sheet created: {sheet_name}")
            else:
                sheet = workbook[sheet_name]
    
            # Append data to the sheet row by row
            for row in df.itertuples(index=False):
                sheet.append(row)
    
            # Freeze the first row and third column for better viewing
            sheet.freeze_panes = 'D2'  # Freeze everything above row 2 and to the left of column C
    
            # Save the updated Excel file
            workbook.save(EXCEL_FILE_PATH)
            log_message(f"Data successfully appended to Excel file: {EXCEL_FILE_PATH}_{sheet_name}")
            

def validate_input(value, min_val=None):
    if value is None or pd.isna(value) or not isinstance(value, (int, float)):
        return None
    if min_val is not None and value < min_val:
        return None
    return value
        
def calculate_individual_scores(pe, dividend_yield, earnings_growth):

    dividend_yield = dividend_yield * 100

    # P/E Ratio Scoring Logic (Lower is better)
    if pe is not None:
        if pe <= 10:
            pe_score = 5
        elif pe <= 20:
            pe_score = 4
        elif pe <= 30:
            pe_score = 3
        elif pe <= 50:
            pe_score = 2
        else:
            pe_score = 1
    else:
        pe_score = 0

    # Dividend Yield Scoring Logic (Higher is better)
    if dividend_yield is not None:
        if dividend_yield > 4:
            dividend_score = 5
        elif dividend_yield > 3:
            dividend_score = 4
        elif dividend_yield > 2:
            dividend_score = 3
        elif dividend_yield > 1:
            dividend_score = 2
        else:
            dividend_score = 1
    else:
        dividend_score = 0

    # Earnings Growth Scoring Logic (Higher is better)
    if earnings_growth is not None:
        if earnings_growth > 20:
            earnings_growth_score = 5
        elif earnings_growth > 10:
            earnings_growth_score = 4
        elif earnings_growth > 5:
            earnings_growth_score = 3
        elif earnings_growth > 0:
            earnings_growth_score = 2
        else:
            earnings_growth_score = 1
    else:
        earnings_growth_score = 0

    # If any of the scores are invalid, return "None"
    if pe_score == 0 or dividend_score == 0 or earnings_growth_score == 0:
        return None

    # Weighted total score (scaled to 1-5)
    total_score = (pe_score * 0.4) + (dividend_score * 0.3) + (earnings_growth_score * 0.3)
    total_score = round(total_score, 1)
    
    # Assign Calculated Recommendation
    if total_score <= 1.5:
        recommendation = "Strong Buy"
    elif total_score <= 2.5:
        recommendation = "Buy"
    elif total_score <= 3.5:
        recommendation = "Hold"
    elif total_score <= 4.5:
        recommendation = "Underperform"
    else:
        recommendation = "Sell"

    return total_score, recommendation
    #return round(total_score, 1)

def analyze_stock_with_profiles(info):
    recommendations = []
    
    try:
        # Extract relevant fields
        beta = info.get('beta', 'N/A')
        pe_ratio = info.get('trailingPE', 'N/A')
        forward_pe = info.get('forwardPE', 'N/A')
        dividend_yield = info.get('dividendYield', 'N/A')
        price_to_book = info.get('priceToBook', 'N/A')
        profit_margins = info.get('profitMargins', 'N/A')
        revenue_growth = info.get('revenueGrowth', 'N/A')
        high_52w = info.get('fiftyTwoWeekHigh', 'N/A')
        low_52w = info.get('fiftyTwoWeekLow', 'N/A')
        recommendation_mean = info.get('recommendationMean', 'N/A')
        current_price = info.get('currentPrice', 'N/A')

        # 1. Conservative Investor (Low Risk, Income-Focused)
        if beta != 'N/A' and beta < 1:
            conservative_reason = "Low Beta (less volatile than the market)"
        else:
            conservative_reason = "High Beta (more volatile)"
        
        if dividend_yield != 'N/A' and dividend_yield > 0.03:
            conservative_reason += "- Pays a good dividend (>3%)"
        
        if price_to_book != 'N/A' and price_to_book < 1:
            conservative_reason += "- Price-to-Book ratio (<1) indicates undervalued assets"
        elif price_to_book != 'N/A' and price_to_book < 2:
            conservative_reason += "- Price-to-Book ratio (<2) indicates potential for growth"
        
        recommendations.append({
            "Cal_Investment_Profile": "Conservative Investor",
            "Cal_Recommendation": "Buy" if dividend_yield != 'N/A' and dividend_yield > 0.03 else "Hold",
            "Cal_Reason": conservative_reason
        })

        # 2. Growth Investor (Focus on High Growth)
        growth_reason = []
        if forward_pe != 'N/A' and forward_pe < 20:
            growth_reason.append("Low Forward P/E (<20) indicates growth potential")
        if revenue_growth != 'N/A' and revenue_growth > 0.1:
            growth_reason.append("Strong Revenue Growth (>10%)")
        if profit_margins != 'N/A' and profit_margins > 0.2:
            growth_reason.append("Highly Profitable with margins > 20%")
        
        if growth_reason:
            recommendations.append({
                "Cal_Investment_Profile": "Growth Investor",
                "Cal_Recommendation": "Buy" if forward_pe != 'N/A' and forward_pe < 20 else "Hold",
                "Cal_Reason": "- ".join(growth_reason)
            })
        else:
            # Add a default entry with None as the reason
            recommendations.append({
                "Cal_Investment_Profile": "Growth Investor",
                "Cal_Recommendation": "None",
                "Cal_Reason": "None"
            })

        # 3. Momentum Investor (Focus on Recent Trends)
        momentum_reason = []
        if high_52w != 'N/A' and low_52w != 'N/A':
            price_position = (current_price - low_52w) / (high_52w - low_52w)
            if price_position > 0.75:
                momentum_reason.append("Trading near its 52-week high (bullish momentum)")
            elif price_position < 0.25:
                momentum_reason.append("Trading near its 52-week low (bearish momentum)")
        else:
            momentum_reason.append("None")
       
        recommendations.append({
            "Cal_Investment_Profile": "Momentum Investor",
            "Cal_Recommendation": "Buy" if price_position > 0.75 else "Hold",
            "Cal_Reason": "- ".join(momentum_reason) if momentum_reason else "No strong momentum"
        })

    except Exception as e:
        recommendations.append({
            "Cal_Investment_Profile": "Error",
            "Cal_Recommendation": "None",
            "Cal_Reason": "None"
        })
    
    return recommendations
    
    
def fetch_and_update_stock_data(symbol, total_symbol):
    try:
        # Read the current row counter
        current_counter = get_current_row_counter()

        log_message(f"Life count: {current_counter} Fetching data for NSE: {symbol} ...")
        stock = yf.Ticker(symbol)
        info = stock.info
        
        # Safely access data with default values
        pe_ratio = info.get('trailingPE', 0)
        dividend_yield = info.get('dividendYield', 0)
        earnings_growth = info.get('earningsQuarterlyGrowth', 0)        

        # Calculate score (assuming the `calculate_individual_scores` function is defined)
        score, score_recommendation = calculate_individual_scores(pe_ratio, dividend_yield, earnings_growth)

        cal_recom = analyze_stock_with_profiles(info)

        current_price = info.get('currentPrice', 'N/A')
        previous_close = info.get('previousClose', 'N/A')

        # Calculate today's growth percentage
        if current_price != 'N/A' and previous_close != 'N/A' and previous_close != 0:
            today_growth_percentage = ((current_price - previous_close) / previous_close) * 100
            today_growth_percentage = round(today_growth_percentage, 2)
        else:
            today_growth_percentage = 'N/A'

        PREVIOUS_DAY_DATETIME = ist_now.strftime('%Y-%m-%d %H:%M:%S')
        # Extract data and include the Previous Day Date
        info_row = [current_counter, PREVIOUS_DAY_DATETIME, symbol] + [info.get(key, '') for key in headers]  + [today_growth_percentage, score, score_recommendation]

        for recom in cal_recom:
            info_row.append(recom.get("Cal_Recommendation", ""))
            info_row.append(recom.get("Cal_Reason", ""))

        # Append data to CSV and Excel
        append_to_csv(info_row, total_symbol)

        # Increment row_insert_order for the next row
        current_counter += 1
        update_row_counter(current_counter)

        return info_row
    except Exception as e:
        log_message(f"Error fetching data for NSE {symbol}: {e}")
        return None

# Add process data
def preprocess_data(csv_file_path):
    """
    Preprocess the CSV file to ensure data types are correct based on the BigQuery schema.
    If incorrect types are detected, log the error and attempt to fix them.
    """

    processed_rows = []
    errors = []

    try:
        with open(csv_file_path, "r") as csv_file:
            reader = csv.DictReader(csv_file)
            for row in reader:
                processed_row = {}
                for key, value in row.items():
                    expected_type = data_type_map.get(key, "STRING")
                    try:
                        if expected_type == "STRING":
                            processed_row[key] = value.strip() if value else ""  # Handle empty strings as None
                        elif expected_type == "INTEGER":
                            processed_row[key] = int(value) if value else 0
                        elif expected_type == "FLOAT":
                            try:
                                # Attempt to convert the value to a float
                                processed_row[key] = float(value)
                            except (ValueError, TypeError):
                                # If conversion fails, set it to None
                                processed_row[key] = None
                        elif expected_type == "DATETIME":
                            processed_row[key] = (
                                datetime.strptime(value, "%Y-%m-%d %H:%M:%S") 
                                if value 
                                else datetime(1990, 1, 1, 0, 0, 0)  # Default to '1990-01-01 00:00:00' if no value is provided
                            )
                        elif expected_type == "DATE":
                            try:
                                # Check if value is a Unix timestamp and convert it
                                if value.isdigit():
                                    processed_row[key] = datetime.fromtimestamp(int(value)).date()
                                else:
                                    # Parse date string in the format "YYYY-MM-DD"
                                    processed_row[key] = datetime.strptime(value, "%Y-%m-%d").date()
                            except Exception:
                                # Handle invalid or missing date values with a default date
                                processed_row[key] = datetime(1990, 1, 1).date()
                        else:  # STRING
                            processed_row[key] = ""
                    except (ValueError, TypeError, KeyError) as ve:
                        errors.append(
                            f"Row {processed_count}, Field '{key}' with value '{value}' failed conversion to {expected_type}: {ve}"
                        )
                        processed_row[key] = ""  # Default to None on error
                                        
                processed_rows.append(processed_row)
    except Exception as e:
        log_message(f"Error reading or processing CSV file: {e}")
    
    # Log errors, if any
    if errors:
        log_message(f"Data type errors detected during preprocessing:\n" + "\n".join(errors))
    
    log_message(f"Preprocessing completed for {len(processed_rows)} rows.")
    return processed_rows

def load_data_to_bigquery():
    """Load data from the preprocessed CSV file into BigQuery."""
    try:
        processed_data = preprocess_data(Daily_CSV_FILE_PATH)
        
        # Write processed data back to a temporary CSV for BigQuery loading
        temp_csv_path = "temp_processed.csv"
        
        # Check if the file exists, and delete it if it does
        if os.path.exists(temp_csv_path):
            os.remove(temp_csv_path)
            log_message(f"Deleted the file before start - {temp_csv_path}.")
            
        with open(temp_csv_path, "w", newline="") as temp_csv:
            writer = csv.DictWriter(temp_csv, fieldnames=processed_data[0].keys())
            writer.writeheader()  # Write headers
            writer.writerows(processed_data)  # Write processed rows
        
        log_message(f"Start to load data to BigQuery from {temp_csv_path}.")

        # Load the processed data into BigQuery
        with open(temp_csv_path, "rb") as csv_file:
            job_config = bigquery.LoadJobConfig(
                source_format=bigquery.SourceFormat.CSV,
                skip_leading_rows=1,  # Skip header row
                write_disposition="WRITE_APPEND",  # Append data schema=schema
                autodetect=False,
                max_bad_records=500,  # Tolerate up to 50 bad rows
               # ignore_unknown_values=True,  # Ignore unexpected columns
            )
            load_job = bq_client.load_table_from_file(
                csv_file, BQ_TABLE, job_config=job_config
            )
            load_job.result()  # Wait for the job to complete
            log_message(f"Data loaded to BigQuery table '{BQ_TABLE}' from {temp_csv_path}.")
            # Check for errors
            if load_job.errors:
                log_message(f"Errors encountered during loading: {load_job.errors}")
            else:
                log_message("Data loaded successfully, no errors.")
            log_message(f"Data successfully loaded to BigQuery from {temp_csv_path}.")
    except Exception as e:
        log_message(f"Error loading data to BigQuery: {e}")


def check_market_holiday(symbol, start_date, end_date):
    """
    Checks if the given date is a market holiday for the given symbol.
    If the data is empty (indicating no market data for that date), 
    it prints 'Today is a market holiday' and exits the script.
    """
    data = yf.download(tickers=symbol, start=start_date, end=end_date, interval="1d", auto_adjust=True, progress=False)
    log_message(f"Checking market holiday for Data for {symbol} on {start_date} to {end_date}")
    log_message(f"Data: {data}")
    
    # Check if the data includes the specified date range
    if data.empty or not (start_date in data.index and end_date in data.index):
        log_message(f"Today is holiday or weekend. No trading data available for {symbol} from {start_date} to {end_date}.")
        sys.exit()  # Exit the script if it's a market holiday
    else:
        log_message("Today is a trading day.")

start_date = ist_now.strftime("%Y-%m-%d")
Tomorrow = ist_now + timedelta(days=1)
 
check_market_holiday("RELIANCE.NS", start_date, Tomorrow)
 
# Process each symbol
processed_count = 0

# Process each symbol
for symbol in symbols:
    processed_count += 1
    fetch_and_update_stock_data(symbol, len(symbols))

    # Add a delay to avoid rate-limiting
    time.sleep(0.7)
    log_message(f"Processed {processed_count}/{len(symbols)} symbols.")


def load_data_to_gsheet(spreadsheet):
    # Google Sheet and worksheet names
    sheet_name = f"NSE_{current_year}"

    # Step 1: Read data from the CSV file
    df = pd.read_csv(consolidated_file_path)
    df = df.fillna("")  # Replace NaN with empty strings for Google Sheets compatibility
    log_message(f"Data loaded from CSV file: {consolidated_file_path} for Google Sheets.")
    # Step 3: Open or create the worksheet
    try:
        worksheet = spreadsheet.worksheet(sheet_name)
        log_message(f"Worksheet {sheet_name} opened successfully.")
    except gspread.WorksheetNotFound:
        # Add extra rows to accommodate more data
        extra_rows = len(df) + 200
        worksheet = spreadsheet.add_worksheet(title=sheet_name, rows=str(extra_rows), cols=str(len(df.columns)))
        log_message(f"Worksheet {sheet_name} created successfully.")

    # Step 4: Fetch existing data from the worksheet
    existing_data = worksheet.get_all_values()
    if existing_data:
        existing_df = pd.DataFrame(existing_data[1:], columns=existing_data[0])  # Convert to DataFrame
        log_message(f"Existing data fetched from worksheet {sheet_name}.")
    else:
        existing_df = pd.DataFrame()  # If the worksheet is empty, start fresh
        log_message(f"No existing data found in worksheet {sheet_name}.")

    # Step 5: Combine the new data with existing data
    if not existing_df.empty:
        # Combine new data with existing data and drop duplicates
        merged_df = pd.concat([existing_df, df], ignore_index=True).drop_duplicates()
        # Ensure unmatched columns from existing_df are retained
        for col in existing_df.columns:
            if col not in merged_df.columns:
                merged_df[col] = existing_df[col]
    else:
        # If the sheet is empty, just use the new data
        merged_df = df

    # Step 6: Write updated data back to the worksheet
    # Prepare data for writing
    data_to_update = [merged_df.columns.tolist()] + merged_df.fillna("").values.tolist()

    # Write data back to the worksheet
    worksheet.update(data_to_update)
    log_message(f"*** Data updated successfully in worksheet {sheet_name}. ***")
    
# BigQuery authentication
bq_client = bigquery.Client.from_service_account_json(SERVICE_ACCOUNT_FILE)

# King file for the year
create_consolidated_csv()

gc = gspread.service_account(filename=SERVICE_ACCOUNT_FILE)
# Open Google Spreadsheet
indian_spreadsheet = gc.open('Indian_Stock_Data')  # Replace with your Google Sheet name

load_data_to_gsheet(indian_spreadsheet)

# Ensure dataset and table exist in BigQuery
ensure_dataset_exists()
ensure_table_exists()

# Load the data into BigQuery from the CSV file
load_data_to_bigquery()

log_message("NSE Daily Script execution completed.")
